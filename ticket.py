import tkinter as tk
from PIL import Image
from PIL import ImageTk, Image
from tkinter import filedialog
from tkinter import *
from tkinter.filedialog import askopenfilename
from pathlib import Path
import csv
from PIL import Image
import os
print("select your EXCEL file:")
tk.Tk().withdraw() # part of the import if you are not using other tkinter functions
fn = askopenfilename()
print("CSV file:", fn)
fn1=fn.replace("/","\\")
print (fn1)

print("Select your folder:")
root = Tk()
root.withdraw()
fs = filedialog.askdirectory()
print("your selected folder:",fs)
fs1=fs.replace("/","\\")
print (fs1)

print("select your main image file:")
tk.Tk().withdraw() # part of the import if you are not using other tkinter functions
pn = askopenfilename()
print("Image file:", pn)
pn1=pn.replace("/","\\")
img1 = Image.open(str(pn1))
n="processed"
parent_dir = str(fs1)+"\\"
path = os.path.join(parent_dir, n)
os.mkdir(path)
print("Directory '% s' created" % n)
print(path)

import openpyxl

# Define variable to load the dataframe
dataframe = openpyxl.load_workbook(str(fn1))

# Define variable to read sheet
dataframe1 = dataframe.active

# Iterate the loop to read the cell values
for row in range(0, dataframe1.max_row):
	for col in dataframe1.iter_cols(1, dataframe1.max_column):
		print(col[row].value+"done...")
		img2 = Image.open(str(fs1)+"\\"+str(col[row].value)+".png")
		img1.paste(img2, (1630, 190))
		img1.save(str(path)+"\\"+str(col[row].value)+"new"+".png")
print("tickets are saved to location: ",path)